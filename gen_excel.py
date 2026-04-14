# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "RAB SITRIA"

BLUE_HDR   = "1F3864"
LIGHT_BLUE = "BDD7EE"
BLUE2      = "D9E1F2"
WHITE      = "FFFFFF"

thin = Side(style="thin", color="000000")
ALL_THIN = Border(top=thin, bottom=thin, left=thin, right=thin)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def put(ws, r, c, val=None, bold=False, h="center", v="center",
        wrap=False, bg=None, fc="000000", sz=9, brd=True, fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = Font(bold=bold, color=fc, size=sz, name="Calibri")
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    if bg:  cell.fill   = fill(bg)
    if brd: cell.border = ALL_THIN
    if fmt: cell.number_format = fmt
    return cell

# Column widths
widths = {1:10, 2:4, 3:48, 4:4, 5:5, 6:3, 7:4, 8:5, 9:3,
          10:4, 11:5, 12:4, 13:5, 14:7, 15:14, 16:16, 17:16}
for c, w in widths.items():
    ws.column_dimensions[get_column_letter(c)].width = w

# ── ROWS 1-3  Headers
r = 1
ws.row_dimensions[r].height = 20
for c in range(1, 18):
    put(ws, r, c, bg=BLUE_HDR, fc=WHITE, bold=True)

ws.merge_cells("A1:A3")
put(ws, 1, 1, "Kode", bold=True, bg=BLUE_HDR, fc=WHITE, sz=10)

ws.merge_cells("B1:B3")
put(ws, 1, 2, "", bg=BLUE_HDR)

ws.merge_cells("C1:C3")
put(ws, 1, 3, "Uraian Suboutput/Komponen/\nSubkomponen/Akun/Detil",
    bold=True, bg=BLUE_HDR, fc=WHITE, sz=10, wrap=True)

ws.merge_cells("D1:Q1")
put(ws, 1, 4, "ANGGARAN", bold=True, bg=BLUE_HDR, fc=WHITE, sz=11)

r = 2
ws.row_dimensions[r].height = 22
ws.merge_cells("D2:M2")
put(ws, 2, 4, "Rincian Perhitungan", bold=True, bg=BLUE_HDR, fc=WHITE)
for c in [14, 15, 16, 17]:
    put(ws, 2, c, bg=BLUE_HDR)
put(ws, 2, 14, "Jumlah",                   bold=True, bg=BLUE_HDR, fc=WHITE)
put(ws, 2, 15, "Harga Satuan",             bold=True, bg=BLUE_HDR, fc=WHITE, wrap=True)
put(ws, 2, 16, "Jumlah",                   bold=True, bg=BLUE_HDR, fc=WHITE)
put(ws, 2, 17, "Waktu Pelaksanaan\nKegiatan", bold=True, bg=BLUE_HDR, fc=WHITE, wrap=True)

r = 3
ws.row_dimensions[r].height = 16
ws.merge_cells("D3:M3")
put(ws, 3, 4, "Perhitungan", bold=True, bg=BLUE_HDR, fc=WHITE)
for c in [1, 2, 3, 14, 15, 16, 17]:
    put(ws, 3, c, bg=BLUE_HDR)

# ── DATA
ITEMS = [
    ("a", "Pengembangan Fitur Statistik dan Visualisasi Data Publikasi Ilmiah (SITRIA)",  4_000_000),
    ("b", "Pengembangan Fitur Galeri Karya dan Inovasi Penelitian",                       2_500_000),
    ("c", "Pengembangan Fitur Peta Kolaborasi dan Roadmap Riset",                         3_000_000),
    ("d", "Pengembangan Fitur Pencarian dan Pencocokan Kepakaran Dosen",                  2_000_000),
    ("e", "Pengembangan Fitur Monitoring Dana, Hibah, dan Data Akreditasi (DTPS)",        3_000_000),
    ("f", "Desain Antarmuka Pengguna dan Pengembangan Infrastruktur Sistem",              3_500_000),
    ("g", "Pengembangan Modul Otomasi Pengambilan Data dan Analisis Kecerdasan Buatan",   5_000_000),
    ("h", "Pengembangan Sistem Pengelolaan Data Dosen, Publikasi, dan Multi-Program Studi",3_000_000),
    ("i", "Pemasangan Sistem, Uji Coba, dan Penyusunan Panduan Pengguna",                 2_000_000),
]
subtotal = sum(x[2] for x in ITEMS)
ppn      = round(subtotal * 0.11)
total    = subtotal + ppn

# ROW 4 — section title
r = 4
ws.row_dimensions[r].height = 18
put(ws, r,  1, "A",   bold=True, bg=LIGHT_BLUE)
put(ws, r,  2, "",    bg=LIGHT_BLUE)
put(ws, r,  3, "PELAKSANAAN GALERI/KATALOG INOVASI (WEB SITRIA)",
    bold=True, h="left", bg=LIGHT_BLUE)
for c in range(4, 16):
    put(ws, r, c, bg=LIGHT_BLUE)
put(ws, r, 16, total, bold=True, h="right", bg=LIGHT_BLUE, fmt="#,##0.00")
put(ws, r, 17, "",    bg=LIGHT_BLUE)

# ROWS 5-13  items a-i
for idx, (letter, uraian, harga) in enumerate(ITEMS):
    r = 5 + idx
    ws.row_dimensions[r].height = 30
    put(ws, r,  1, "522131",       h="center")
    put(ws, r,  2, letter,         h="center")
    put(ws, r,  3, uraian,         h="left", v="center", wrap=True)
    put(ws, r,  4, 1,              h="center")
    put(ws, r,  5, "PKT",          h="center")
    put(ws, r,  6, "x",            h="center")
    put(ws, r,  7, 10,             h="center")
    put(ws, r,  8, "KEG",          h="center")
    put(ws, r,  9, "x",            h="center")
    put(ws, r, 10, 10,             h="center")
    put(ws, r, 11, "KALI",         h="center")
    put(ws, r, 12, 1,              h="center")
    put(ws, r, 13, "PKT",          h="center")
    put(ws, r, 14, 1,              h="center", fmt="#,##0")
    put(ws, r, 15, harga,          h="right",  fmt="#,##0")
    put(ws, r, 16, harga,          h="right",  fmt="#,##0.00")
    put(ws, r, 17, "Maret - April",h="center")

# ROW 14  PPN
r = 14
ws.row_dimensions[r].height = 16
put(ws, r,  1, "522131",   h="center")
put(ws, r,  2, "j",        h="center")
put(ws, r,  3, "PPN 11%",  h="left")
for c in range(4, 16):
    put(ws, r, c, "")
put(ws, r, 16, ppn,        h="right", fmt="#,##0.00")
put(ws, r, 17, "Maret - April", h="center")

# ROW 15  TOTAL
r = 15
ws.row_dimensions[r].height = 18
for c in range(1, 18):
    put(ws, r, c, bg=BLUE2)
put(ws, r,  3, "TOTAL", bold=True, h="right", bg=BLUE2, sz=10)
put(ws, r, 16, total,   bold=True, h="right", bg=BLUE2, fmt="#,##0.00", sz=10)

out = "d:/Github-ADL/presentasi-laboratorium-inovasi-digital/RAB_SITRIA_2026.xlsx"
wb.save(out)
print("Saved: " + out)
print("Subtotal : Rp " + f"{subtotal:,}")
print("PPN 11%  : Rp " + f"{ppn:,}")
print("Total    : Rp " + f"{total:,}")
