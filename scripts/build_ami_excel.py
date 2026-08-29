# -*- coding: utf-8 -*-
"""
Build Excel Pemetaan Judul Penelitian -> Roadmap FSTI (bukti AMI)
=================================================================
Menghasilkan workbook siap cetak/serah untuk butir AMI:
"Apakah program studi memiliki pemetaan judul penelitian terhadap
roadmap penelitian jurusan / ITK?"

Input : src/data/prodi/{slug}/roadmap_mapping.json
        (dihasilkan oleh scripts/build_roadmap_mapping.py)
Output: Pemetaan_Roadmap_Penelitian_{PRODI}.xlsx di root project

Usage:
  python scripts/build_ami_excel.py sistem-informasi
"""

import io
import json
import os
import sys
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PRODI_DIR = os.path.join(BASE_DIR, 'src', 'data', 'prodi')

# Palet: biru institusi + satu warna per pilar
NAVY = '1F3864'
GREY_HDR = 'D9E1F2'
PILLAR_FILL = {'P1': 'DCE6F1', 'P2': 'E6DFF0', 'P3': 'DDEEE4', 'P4': 'FAECD9', 'NA': 'EDEDED'}
PILLAR_FONT = {'P1': '2C5C8F', 'P2': '6B4F9C', 'P3': '2F7F5E', 'P4': 'A4661F', 'NA': '767676'}
SHORT = {'P1': 'Smart Governance', 'P2': 'Smart Education',
         'P3': 'Smart Living', 'P4': 'Smart Grid & Energi', 'NA': 'Di luar pilar'}
ORDER = ['P1', 'P2', 'P3', 'P4']

thin = Side(style='thin', color='BFBFBF')
BOX = Border(top=thin, bottom=thin, left=thin, right=thin)


def fill(hexcolor):
    return PatternFill('solid', fgColor=hexcolor)


def idn(iso):
    bulan = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni', 'Juli',
             'Agustus', 'September', 'Oktober', 'November', 'Desember']
    y, m, d = iso[:10].split('-')
    return '%d %s %s' % (int(d), bulan[int(m) - 1], y)


def header_row(ws, row, labels, widths=None):
    for i, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = fill(NAVY)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BOX
    ws.row_dimensions[row].height = 30
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def title_block(ws, judul, subjudul, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=judul)
    c.font = Font(bold=True, size=14, color=NAVY)
    c.alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 24
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    c = ws.cell(row=2, column=1, value=subjudul)
    c.font = Font(size=9, italic=True, color='595959')
    ws.row_dimensions[3].height = 6


def build(slug):
    src = os.path.join(PRODI_DIR, slug, 'roadmap_mapping.json')
    if not os.path.exists(src):
        print('  [error] %s belum ada. Jalankan build_roadmap_mapping.py dulu.' % src)
        return
    d = json.load(io.open(src, encoding='utf-8'))
    items, P, S = d['items'], d['pillars'], d['summary']
    prodi = d['metadata']['prodi']
    src_date = idn(d['metadata']['sourceGeneratedAt'])
    gen_date = idn(d['metadata']['generatedAt'])

    years = sorted({x['year'] for x in items})
    byp, fund, yearmat = defaultdict(list), defaultdict(int), defaultdict(Counter)
    for x in items:
        k = x['pillar'] or 'NA'
        byp[k].append(x)
        fund[k] += x['fundingAmount']
        yearmat[k][x['year']] += 1
    for k in byp:
        byp[k].sort(key=lambda r: -int(r['year'] or 0))
    total = S['totalUniqueResearch']

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ 1. RINGKASAN
    ws = wb.active
    ws.title = 'Ringkasan'
    title_block(ws, 'PEMETAAN JUDUL PENELITIAN TERHADAP ROADMAP PENELITIAN FSTI ITK',
                'Dokumen bukti Audit Mutu Internal (AMI) - Kriteria Penelitian', 6)
    for w, col in zip([26, 34, 14, 14, 20, 16], 'ABCDEF'):
        ws.column_dimensions[col].width = w

    r = 4
    ws.cell(row=r, column=1, value='PERTANYAAN AMI').font = Font(bold=True, size=10, color=NAVY)
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(row=r, column=1,
                value='Apakah program studi memiliki pemetaan judul penelitian '
                      'terhadap roadmap penelitian jurusan / ITK?')
    c.font = Font(italic=True, size=11)
    c.alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[r].height = 20

    r += 2
    ws.cell(row=r, column=1, value='JAWABAN').font = Font(bold=True, size=10, color=NAVY)
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=6)
    c = ws.cell(row=r, column=1, value=(
        'YA, TERSEDIA. Seluruh %d judul penelitian dosen tetap Prodi %s periode %s-%s telah dipetakan '
        'ke 4 Pilar Strategis Roadmap Penelitian, PkM, dan Renstra FSTI ITK 2025-2029. Sebanyak %d judul '
        '(%s%%) terpetakan pada salah satu pilar; %d judul merupakan riset fondasi keilmuan yang menopang '
        'klaster SDM peneliti FSTI. Pemetaan dihasilkan otomatis dari data SINTA Kemendiktisaintek melalui '
        'skrip scripts/build_roadmap_mapping.py sehingga dapat direproduksi dan diaudit ulang.'
        % (total, prodi, years[0], years[-1], S['mapped'], S['coveragePercent'], S['unmapped'])))
    c.font = Font(size=10)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    c.fill = fill('E2EFDA')
    c.border = BOX

    r += 4
    ws.cell(row=r, column=1, value='REKAPITULASI').font = Font(bold=True, size=10, color=NAVY)
    r += 1
    header_row(ws, r, ['Pilar Strategis Roadmap FSTI', 'Fokus', 'Jumlah Judul',
                       'Proporsi', 'Total Pendanaan (Rp)', 'Rentang Tahun'])
    ws.freeze_panes = None
    r += 1
    first_data = r
    for p in ORDER + ['NA']:
        n = len(byp[p])
        nama = ('Pilar %s - %s' % (p[1], P[p]['name'])) if p != 'NA' else 'Di luar keempat pilar (riset fondasi)'
        tahun = sorted({x['year'] for x in byp[p]})
        vals = [nama, SHORT[p], n, n / total, fund[p],
                ('%s-%s' % (tahun[0], tahun[-1])) if len(tahun) > 1 else (tahun[0] if tahun else '-')]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BOX
            c.font = Font(size=10, bold=(i == 1), color=PILLAR_FONT[p] if i <= 2 else '000000')
            c.fill = fill(PILLAR_FILL[p])
            if i == 1:
                c.alignment = Alignment(wrap_text=True, vertical='center')
            if i == 3:
                c.alignment = Alignment(horizontal='center')
            if i == 4:
                c.number_format = '0.0%'
                c.alignment = Alignment(horizontal='center')
            if i == 5:
                c.number_format = '#,##0'
        ws.row_dimensions[r].height = 28
        r += 1
    for i, v in enumerate(['TOTAL', '', total, 1.0, sum(fund.values()), '%s-%s' % (years[0], years[-1])], start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = Font(bold=True, size=10)
        c.fill = fill(GREY_HDR)
        c.border = BOX
        if i in (3, 4):
            c.alignment = Alignment(horizontal='center')
        if i == 4:
            c.number_format = '0.0%'
        if i == 5:
            c.number_format = '#,##0'

    r += 3
    ws.cell(row=r, column=1, value='SUMBER DATA DAN BERKAS BUKTI').font = Font(bold=True, size=10, color=NAVY)
    r += 1
    bukti = [
        ('Roadmap_FSTI.pdf', 'Dokumen acuan - Roadmap Penelitian, PkM & Renstra FSTI ITK 2025-2029'),
        ('src/data/prodi/%s/roadmap_mapping.json' % slug, 'Hasil pemetaan (%d judul + skor kecocokan)' % total),
        ('scripts/build_roadmap_mapping.py', 'Skrip pemetaan - aturan klasifikasi terbuka'),
        ('src/data/prodi/%s/sinta_data.json' % slug, 'Data sumber - tarikan SINTA per %s' % src_date),
        ('scripts/sinta_scraper.py', 'Skrip penarikan data SINTA Kemendiktisaintek'),
    ]
    header_row(ws, r, ['Berkas', 'Keterangan', '', '', '', ''])
    ws.freeze_panes = None
    r += 1
    for f, ket in bukti:
        ws.cell(row=r, column=1, value=f).font = Font(size=9, name='Consolas')
        ws.cell(row=r, column=1).border = BOX
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        c = ws.cell(row=r, column=2, value=ket)
        c.font = Font(size=9)
        c.border = BOX
        r += 1

    r += 2
    ws.cell(row=r, column=1, value='Data SINTA ditarik %s | Pemetaan disusun %s' % (src_date, gen_date)).font = \
        Font(size=8, italic=True, color='808080')
    r += 1
    ws.cell(row=r, column=1, value='Program Studi %s - Fakultas Sains dan Teknologi Informasi - Institut Teknologi Kalimantan' % prodi).font = \
        Font(size=8, italic=True, color='808080')

    # ------------------------------------------------------- 2. PEMETAAN JUDUL
    ws = wb.create_sheet('Pemetaan Judul')
    title_block(ws, 'DAFTAR PEMETAAN JUDUL PENELITIAN KE PILAR ROADMAP FSTI',
                'Prodi %s | %d judul unik | Sumber: SINTA per %s' % (prodi, total, src_date), 8)
    header_row(ws, 4,
               ['No', 'Tahun', 'Judul Penelitian', 'Ketua Peneliti', 'Skema Hibah',
                'Pendanaan (Rp)', 'Pilar Utama', 'Pilar Pendukung'],
               [5, 8, 72, 26, 14, 16, 30, 14])
    r = 5
    no = 0
    for p in ORDER + ['NA']:
        nama = ('PILAR %s - %s (%d judul)' % (p[1], P[p]['name'].upper(), len(byp[p]))) if p != 'NA' \
            else 'DI LUAR KEEMPAT PILAR - RISET FONDASI KEILMUAN (%d judul)' % len(byp[p])
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        c = ws.cell(row=r, column=1, value=nama)
        c.font = Font(bold=True, size=10, color=PILLAR_FONT[p])
        c.fill = fill(PILLAR_FILL[p])
        c.border = BOX
        c.alignment = Alignment(vertical='center')
        ws.row_dimensions[r].height = 20
        r += 1
        for x in byp[p]:
            no += 1
            anggota = len(x['lecturers'])
            vals = [no, x['year'], x['title'],
                    (x.get('leader') or '-') + (' (+%d dosen prodi)' % (anggota - 1) if anggota > 1 else ''),
                    x.get('grantCategory') or '-', x['fundingAmount'],
                    ('Pilar %s - %s' % (p[1], SHORT[p])) if p != 'NA' else 'Di luar pilar',
                    ', '.join(x['supportingPillars']) or '-']
            for i, v in enumerate(vals, start=1):
                c = ws.cell(row=r, column=i, value=v)
                c.border = BOX
                c.font = Font(size=9)
                c.alignment = Alignment(vertical='top',
                                        wrap_text=(i in (3, 4)),
                                        horizontal='center' if i in (1, 2, 5, 8) else 'left')
                if i == 6:
                    c.number_format = '#,##0'
                    c.alignment = Alignment(horizontal='right', vertical='top')
                if i == 7:
                    c.font = Font(size=9, color=PILLAR_FONT[p])
            ws.row_dimensions[r].height = 28
            r += 1
    ws.auto_filter.ref = 'A4:H%d' % (r - 1)

    # --------------------------------------------------------- 3. REKAP TAHUN
    ws = wb.create_sheet('Rekap per Tahun')
    title_block(ws, 'SEBARAN JUDUL PENELITIAN PER TAHUN DAN PILAR',
                'Menunjukkan konsistensi arah riset prodi sebelum dan sesudah roadmap FSTI 2025-2029 ditetapkan', 3)
    header_row(ws, 4, ['Pilar Strategis'] + list(years) + ['Total'],
               [34] + [8] * len(years) + [10])
    r = 5
    for p in ORDER + ['NA']:
        nama = ('Pilar %s - %s' % (p[1], SHORT[p])) if p != 'NA' else 'Di luar pilar'
        c = ws.cell(row=r, column=1, value=nama)
        c.font = Font(size=10, bold=True, color=PILLAR_FONT[p])
        c.fill = fill(PILLAR_FILL[p])
        c.border = BOX
        for j, y in enumerate(years, start=2):
            n = yearmat[p][y]
            c = ws.cell(row=r, column=j, value=n if n else None)
            c.alignment = Alignment(horizontal='center')
            c.border = BOX
            c.font = Font(size=10)
            if n:
                c.fill = fill(PILLAR_FILL[p])
        c = ws.cell(row=r, column=len(years) + 2, value=len(byp[p]))
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal='center')
        c.border = BOX
        r += 1
    c = ws.cell(row=r, column=1, value='TOTAL')
    c.font = Font(bold=True, size=10)
    c.fill = fill(GREY_HDR)
    c.border = BOX
    for j, y in enumerate(years, start=2):
        c = ws.cell(row=r, column=j, value=sum(yearmat[p][y] for p in ORDER + ['NA']))
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal='center')
        c.fill = fill(GREY_HDR)
        c.border = BOX
    c = ws.cell(row=r, column=len(years) + 2, value=total)
    c.font = Font(bold=True, size=10)
    c.alignment = Alignment(horizontal='center')
    c.fill = fill(GREY_HDR)
    c.border = BOX

    # ------------------------------------------------------ 4. ACUAN PILAR
    ws = wb.create_sheet('Acuan Pilar Roadmap')
    title_block(ws, 'ACUAN: EMPAT PILAR STRATEGIS ROADMAP FSTI ITK 2025-2029',
                'Sumber: Roadmap_FSTI.pdf Bab 4 - Pilar Strategis Integratif Riset & Abdimas', 4)
    header_row(ws, 4, ['Pilar', 'Nama Pilar', 'Proyek Unggulan Fakultas',
                       'Peran Prodi %s dalam Pilar' % prodi], [8, 38, 30, 56])
    r = 5
    for p in ORDER:
        vals = [p, P[p]['name'], P[p]['flagship'], P[p]['siRole']]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BOX
            c.font = Font(size=10, bold=(i == 1), color=PILLAR_FONT[p] if i <= 2 else '000000')
            c.fill = fill(PILLAR_FILL[p]) if i <= 2 else fill('FFFFFF')
            c.alignment = Alignment(wrap_text=True, vertical='top',
                                    horizontal='center' if i == 1 else 'left')
        ws.row_dimensions[r].height = 48
        r += 1

    # ------------------------------------------------- 5. TEMUAN & TINDAK LANJUT
    ws = wb.create_sheet('Temuan & Tindak Lanjut')
    title_block(ws, 'TEMUAN EVALUASI DIRI DAN RENCANA TINDAK LANJUT',
                'Hasil analisis pemetaan judul penelitian terhadap roadmap FSTI', 3)
    header_row(ws, 4, ['Aspek', 'Temuan', 'Rencana Tindak Lanjut'], [26, 62, 58])
    temuan = [
        ('Kesesuaian dengan mandat prodi',
         '%d dari %d judul (%.0f%%) berada pada Pilar 1 Smart Governance dan Transformasi Digital Perkotaan. '
         'Selaras dengan peran Prodi %s dalam roadmap FSTI, yaitu perancangan enterprise architecture dan '
         'integrasi layanan e-Government, serta menopang proyek unggulan GovTech Nusantara.'
         % (len(byp['P1']), total, len(byp['P1']) / total * 100, prodi),
         'Dipertahankan. Diarahkan pada peningkatan luaran berupa Hak Cipta Perangkat Lunak sesuai KPI '
         'Roadmap FSTI (minimal 1 HKI per tahun per kelompok riset).'),
        ('Cakupan pilar',
         'Keempat pilar roadmap tercakup, namun Pilar 4 Smart Grid, Energi Cerdas, dan Mobilitas '
         'Berkelanjutan baru disentuh %d judul (2024-2025), seluruhnya berupa sistem monitoring energi.'
         % len(byp['P4']),
         'Mendorong kolaborasi dengan Prodi Teknik Elektro pada sisi antarmuka dan platform data untuk '
         'microgrid serta EV Charging Station pada skema hibah 2027.'),
        ('Judul di luar pilar',
         '%d judul (algoritma metaheuristik 2024 dan adaptive mesh dinamika vorteks 2023) merupakan riset '
         'fondasi keilmuan komputasi, bukan penyimpangan dari roadmap. Keduanya menopang klaster '
         '"Optimasi Dinamik & Teori Graf" pada peta klasterisasi SDM peneliti FSTI.' % S['unmapped'],
         'Diarahkan sebagai metode pendukung salah satu pilar pada usulan penelitian berikutnya agar '
         'kontribusinya terhadap roadmap terbaca eksplisit pada judul.'),
        ('Pemutakhiran pemetaan',
         'Pemetaan dihasilkan skrip dari data SINTA, bukan pendataan manual, sehingga tidak menjadi usang '
         'dan konsisten dengan data resmi kementerian.',
         'Prosedur rutin: jalankan scripts/sinta_scraper.py, lalu scripts/build_roadmap_mapping.py dan '
         'scripts/build_ami_excel.py setiap awal semester.'),
    ]
    r = 5
    for aspek, t, tl in temuan:
        for i, v in enumerate([aspek, t, tl], start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = BOX
            c.font = Font(size=9, bold=(i == 1))
            c.alignment = Alignment(wrap_text=True, vertical='top')
            if i == 1:
                c.fill = fill(GREY_HDR)
        ws.row_dimensions[r].height = 72
        r += 1

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False

    nama_file = 'Pemetaan_Roadmap_Penelitian_%s.xlsx' % prodi.replace(' ', '_')
    dest = os.path.join(BASE_DIR, nama_file)
    wb.save(dest)
    print('  Tersimpan: %s' % dest)
    print('  %d judul | %d terpetakan (%s%%) | P1=%d P2=%d P3=%d P4=%d luar=%d'
          % (total, S['mapped'], S['coveragePercent'], len(byp['P1']), len(byp['P2']),
             len(byp['P3']), len(byp['P4']), len(byp['NA'])))


def main():
    for slug in (sys.argv[1:] or ['sistem-informasi']):
        build(slug)


if __name__ == '__main__':
    main()
